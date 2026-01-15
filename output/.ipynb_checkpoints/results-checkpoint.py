"""
Export utilities for RICE13_FS.

Goals:
- Excel layout (country sheets + global; Stability sheet).
- Single Source of Truth (SSOT):
    * Payoffs: payoff_row_discounted(...) everywhere.
    * Stability: evaluate_stability_for_mask(...) everywhere.
"""

from __future__ import annotations

import logging
from pathlib import Path
from io import BytesIO
from datetime import datetime
from dataclasses import dataclass
from functools import reduce
from typing import Any, Dict, Iterable, List, Tuple
from types import SimpleNamespace
import pandas as pd
import numpy as np
import yaml
import matplotlib.pyplot as plt
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

from RICE13_FS.common.utils import (
    safe_div, years_from_periods, 
    build_solution_spec_id, normalize_exogenous_S,  build_config_fingerprint, _S_solution_to_df,
    )
from RICE13_FS.analysis.solver import _read_exog_S_csv
from RICE13_FS.output.coalition_store import CoalitionStore


from RICE13_FS.solve.coalition import (
    parse_coalition_spec,              # requires regions=countries
    list_internal_neighbors,
    list_external_neighbors,
    coalition_vec_to_member_string,    # readable names for filenames
)

logger = logging.getLogger(__name__)

# --- Helper: pretty, safe config for the Excel "config" sheet
def _config_for_excel(cfg: Dict[str, Any]) -> Dict[str, Any]:
    """
    Keep only human-friendly scalars; summarize bulky objects so the sheet
    stays readable. This does NOT affect any runtime behavior—only display.

    Small, scalar-valued mappings (e.g. a 'tags' dict) are preserved as-is
    so they can be round-tripped via YAML in the Excel 'config' sheet.
    """
    import pandas as pd
    pretty: Dict[str, Any] = {}
    for k, v in (cfg or {}).items():
        if isinstance(v, (str, int, float, bool)) or v is None:
            pretty[k] = v
        elif isinstance(v, dict):
            # Preserve small scalar-valued dicts (e.g. tags) for readability.
            # Larger / nested structures are summarized by type name.
            try:
                if all(isinstance(val, (str, int, float, bool)) or val is None for val in v.values()):
                    pretty[k] = v
                else:
                    pretty[k] = f"[{type(v).__name__}]"
            except Exception:
                pretty[k] = f"[{type(v).__name__}]"
        elif isinstance(v, pd.DataFrame):
            pretty[k] = f"[DataFrame {v.shape[0]}×{v.shape[1]}]"
        elif isinstance(v, pd.Series):
            pretty[k] = f"[Series len={v.shape[0]}]"
        else:
            pretty[k] = f"[{type(v).__name__}]"
    return pretty


def _scenario_tag_for_run(
    run_name: str,
    util_key: str,
    solution: Dict[str, Any],
    config: Dict[str, Any] | None,
    disc_flag: bool,
) -> str:
    """
    Build a human-friendly scenario tag of the form

        <runmode>_<flavor>_<negishi>[_<disc>]

    where:
      - runmode: 'bau', 'planner', 'nash', ... (derived from `run_name`)
      - flavor:  'crra' or 'fs' (from solution['utility'])
      - negishi: 'N' or 'noN'  (from solution/config)
      - disc:    'disc' or 'nodisc' for FS only (from config/fs_disc_mode)

    Special case:
      - For Nash runs, Negishi has no meaning, so we *omit* the `_N` / `_noN`
        part and only use `<runmode>_<flavor>[_<disc>]`.

    This is only for display in the Excel 'config' sheet (and downstream
    pyam exports); it does not affect filenames or runtime behavior.
    """
    # Run mode from the first token of the name (e.g. 'planner_crra' -> 'planner')
    runmode = run_name.split("_", 1)[0] if run_name else "run"

    # Utility flavor: use util_key if provided, otherwise fall back to solution
    util = (util_key or str(solution.get("utility", ""))).lower().strip()
    if not util:
        # BAU or unknown utility: fall back to just the run name/mode
        return run_name or runmode

    # Negishi status from solution or config
    neg = ("negishi_weights" in solution) or bool((config or {}).get("negishi_use", False))
    neg_tag = "N" if neg else "noN"

    # FS discounting alignment: disc vs nodisc
    disc_tag = ""
    if util == "fs":
        disc_tag = "disc" if disc_flag else "nodisc"

    # Bau: scenario is just bau
    if runmode.startswith("bau"):
        return f"{runmode}"

    # Nash: drop Negishi tag entirely (no '_N' or '_noN')
    if runmode.startswith("nash"):
        if util == "fs" and disc_tag:
            return f"{runmode}_{util}_{disc_tag}"
        return f"{runmode}_{util}"

    # Planner: include Negishi tag
    if util == "fs" and disc_tag:
        return f"{runmode}_{util}_{neg_tag}_{disc_tag}"
    return f"{runmode}_{util}_{neg_tag}"

# ---------------------------------------------------------------------------
# Row order + friendly labels (legacy layout preserved)
# ---------------------------------------------------------------------------

REGION_SPEC = {
    # Utilities & discounting
    "U": "Utility (U)",
    "U_pc": "Utility per capita (U/L)",
    "disc": "Discount factor",

    # FS behavior summaries (filled only for FS runs)
    "FS_envy_avg": "FS envy",
    "FS_guilt_avg": "FS guilt",

    # Demography & macro
    "L": "Population (L)",
    "K": "Capital (K)",
    "K_pc": "Capital per capita (K/L)",
    "S": "Savings rate (S)",
    "Q": "Gross Output (Q)",
    "D": "Climate Damage (D)",
    "Y": "Net Output (Y)",
    "Y_pc": "Net Output per capita (Y/L)",
    "C": "Consumption (C)",
    "C_pc": "Consumption per capita (C/L)",
    "I": "Investment (I)",

    # Emissions & policy
    "E_ind": "Industrial Emissions",
    "E": "Emissions (E)",
    "AB": "Abatement expenditure (AB)",
#    "SCC_welfare": "SCC in Welfare terms",
    "SCC_money": "SCC in regional (k$/tC)",
    "carbon_tax": "Carbon tax (k$/tC)",
    "mu": "Abatement rate (μ)",
}

GLOBAL_SPEC = {
    "T_at": "Atmospheric temperature increase",
    "T_lo": "Lower ocean temperature increase",
    "F": "Increase in radiative forcing (W/m²)",

    "E_tot": "Total carbon emissions (GtC per year)",
    "M_at": "Atmospheric carbon stock (GtC)",
    "M_up": "Upper ocean & biosphere carbon stock (GtC)",
    "M_lo": "Deep ocean carbon stock (GtC)",

    "slr": "Sea level rise (total)",
    "slr_TE": "SLR – thermal expansion",

    "gsic_melt": "GSIC melt (flow)",
    "gsic_cum": "GSIC cumulative melt",
    "gsic_remain": "GSIC remaining",

    "gis_melt": "Greenland melt (flow)",
    "gis_cum": "Greenland cumulative melt",
    "gis_remain": "Greenland ice remaining",

    "ais_melt": "Antarctic melt (flow)",
    "ais_cum": "Antarctic cumulative melt",
    "ais_remain": "Antarctic ice remaining",
    
    "SCC_global_money_pc": "SCC global (avg MU) (k$/tC)",
    # Inequality diagnostics (global, representative agents, per-capita consumption)
    "gini": "Global Gini (C/L, representative agents)",
    "atkinson": "Global Atkinson (1.5, C/L, representative agents)",
}

# ---------------------------------------------------------------------------
# SSOT: Stability on discounted payoffs
# ---------------------------------------------------------------------------

@dataclass
class StabilityResult:
    internally_stable: bool
    externally_stable: bool
    fully_stable: bool
    leavers_idx: List[int]
    joiners_idx: List[int]
    i_missing_names: List[str]
    e_missing_names: List[str]

def _mask_of(vec: List[int]) -> str:
    return "".join("1" if int(b) else "0" for b in vec)


def evaluate_stability_for_mask(
    base_vec: List[int],
    *,
    countries: List[str],
    payoff_by_mask: Dict[str, List[float]],
    eps: float = 1e-3,
) -> StabilityResult:
    r"""
    Single source of truth for coalition stability:
    - Internal: members might leave if their payoff increases in S\{i}.
    - External: outsiders might join if their payoff increases in S∪{i}.
    Special-cases:
    - GRAND: no outsiders → externally stable by construction (outsider set empty).
    - SINGLETONS: no internal neighbors we need to check (don't require 000).
    - Compares only against coalitions present in `payoff_by_mask`.
    """
    # Normalize & basic checks
    n = len(countries)
    base_vec = [1 if int(b) else 0 for b in base_vec]
    if len(base_vec) != n:
        raise ValueError(f"base_vec length {len(base_vec)} != len(countries) {n}")

    base_mask = _mask_of(base_vec)
    base_row = payoff_by_mask.get(base_mask)
    if base_row is None:
        raise KeyError(f"Base coalition {base_mask} missing from payoff map.")
    base_pay = base_row
    size = sum(base_vec)

    # ---------- INTERNAL (members leaving)
    leavers_idx: List[int] = []
    i_missing_names: List[str] = []
    if size <= 1:
        # Singletons have no internal neighbors to check (do not require 000)
        internally_stable = True
    else:
        internally_stable = True
        for i, bit in enumerate(base_vec):
            if bit != 1:
                continue
            neigh = base_vec.copy()
            neigh[i] = 0
            nm = _mask_of(neigh)
            row = payoff_by_mask.get(nm)
            if row is None:
                # Missing neighbor: record a warning but do not penalize the stability flag.
                i_missing_names.append(str(countries[i]))
                continue
            delta = float(row[i]) - float(base_pay[i])
            if delta > eps:
                internally_stable = False
                leavers_idx.append(i)

    # ---------- EXTERNAL (outsiders joining)
    joiners_idx: List[int] = []
    e_missing_names: List[str] = []
    outsiders = [i for i, b in enumerate(base_vec) if b == 0]
    if not outsiders:
        # GRAND: no outsiders → externally stable
        externally_stable = True
    else:
        externally_stable = True
        for i in outsiders:
            join = base_vec.copy()
            join[i] = 1
            jm = _mask_of(join)
            row = payoff_by_mask.get(jm)
            if row is None:
                # Missing neighbor: record a warning but do not penalize the stability flag.
                e_missing_names.append(str(countries[i]))
                continue
            delta = float(row[i]) - float(base_pay[i])
            if delta > eps:
                externally_stable = False
                joiners_idx.append(i)

    return StabilityResult(
        internally_stable=internally_stable,
        externally_stable=externally_stable,
        fully_stable=(internally_stable and externally_stable),
        leavers_idx=leavers_idx,
        joiners_idx=joiners_idx,
        i_missing_names=i_missing_names,
        e_missing_names=e_missing_names,
    )


# ---- Debug helpers (exact utilities & deltas) ----

def _debug_dump_internal(base_vec, countries, payoff_by_mask, *, eps: float = 1e-3):
    if not logger.isEnabledFor(logging.DEBUG):
        return
    base_mask = _mask_of(base_vec)
    base_pay = payoff_by_mask.get(base_mask)
    if base_pay is None:
        return
    members = [countries[i] for i, b in enumerate(base_vec) if b]
    logger.debug("[STAB] Base=%s members=%s", base_mask, members)
    missing = []
    improving = []
    for i, b in enumerate(base_vec):
        if not b:
            continue
        neigh_vec = list(base_vec); neigh_vec[i] = 0
        neigh_mask = _mask_of(neigh_vec)
        row = payoff_by_mask.get(neigh_mask)
        if row is None:
            missing.append((countries[i], neigh_mask))
            logger.debug("  leave %-8s mask=%s  -> MISSING", countries[i], neigh_mask)
            continue
        base_i = float(base_pay[i]); neigh_i = float(row[i])
        delta = neigh_i - base_i
        improves = (delta > eps)
        logger.debug("  leave %-8s mask=%s  base=% .6e  neigh=% .6e  Δ=% .6e%s",
                     countries[i], neigh_mask, base_i, neigh_i, delta,
                     "  (IMPROVES)" if improves else "")
        if improves:
            improving.append((countries[i], delta))
    if missing:
        logger.debug("  missing %d internal neighbors", len(missing))
    if improving:
        improving.sort(key=lambda x: x[1], reverse=True)
        logger.debug("  top improvers: %s", improving[:5])

def _debug_dump_external(base_vec, countries, payoff_by_mask, *, eps: float = 1e-3):
    if not logger.isEnabledFor(logging.DEBUG):
        return
    base_mask = _mask_of(base_vec)
    base_pay = payoff_by_mask.get(base_mask)
    if base_pay is None:
        return
    outsiders = [i for i, b in enumerate(base_vec) if not b]
    logger.debug("[STAB] External check base=%s outsiders=%s",
                 base_mask, [countries[i] for i in outsiders])
    for i in outsiders:
        join_vec = list(base_vec); join_vec[i] = 1
        join_mask = _mask_of(join_vec)
        row = payoff_by_mask.get(join_mask)
        if row is None:
            logger.debug("  join  %-8s mask=%s  -> MISSING", countries[i], join_mask)
            continue
        base_i = float(base_pay[i]); join_i = float(row[i])
        delta = join_i - base_i
        wants = (delta > eps)
        logger.debug("  join  %-8s mask=%s  base=% .6e  join = % .6e  Δ=% .6e%s",
                     countries[i], join_mask, base_i, join_i, delta,
                     "  (WANTS TO JOIN)" if wants else "")

# ---------------------------------------------------------------------------
# Excel shaping (legacy layout preserved)
# ---------------------------------------------------------------------------

def _apply_spec(df: pd.DataFrame, spec: dict) -> pd.DataFrame:
    if not spec:
        return df
    preferred = [k for k in spec.keys() if k in df.index]
    rest = [r for r in df.index if r not in spec]
    df = df.loc[preferred + rest]
    df.rename(index=spec, inplace=True)
    return df

def output_format(
    countries: List[str],
    out_unformat: Dict[str, Any],
    periods: List[int],
    L: pd.DataFrame | None = None,
    *,
    tstep: int,
    base_year: int,
) -> Dict[str, pd.DataFrame]:
    periods = sorted(dict.fromkeys(periods))
    years = years_from_periods(periods, base_year, tstep=tstep)

    # dict-like variables in solver output
    var_l = [k for k in out_unformat if isinstance(out_unformat[k], dict)]

    # Decide region-indexed vs global
    region_vars, global_vars = [], []
    for k in var_l:
        keys = list(out_unformat[k].keys())
        if not keys:
            continue
        if isinstance(keys[0], tuple) and len(keys[0]) == 2:
            region_vars.append(k)
        else:
            global_vars.append(k)

    out_form: Dict[str, pd.DataFrame] = {
        i: pd.DataFrame(data=np.nan, index=region_vars, columns=years) for i in countries
    }
    out_form["global"] = pd.DataFrame(data=np.nan, index=global_vars, columns=years)

    # Fill values
    idx = []
    for j in var_l:
        keys = list(out_unformat[j].keys())
        if not keys:
            continue
        if isinstance(keys[0], tuple) and len(keys[0]) > 2:
            continue  # ignore higher-dim tensors
        if isinstance(keys[0], tuple):
            for i in countries:
                for k_idx, k in enumerate(periods):
                    idx.append((i, j, k, k_idx))
        else:
            for k_idx, k in enumerate(periods):
                idx.append(("global", j, k, k_idx))

    for i_ in idx:
        if i_[0] != "global":
            value = out_unformat[i_[1]].get((i_[0], i_[2]), np.nan)
            out_form[i_[0]].at[i_[1], years[i_[3]]] = value
        else:
            value = out_unformat[i_[1]].get(i_[2], np.nan)
            out_form[i_[0]].at[i_[1], years[i_[3]]] = value

    # FS averages only for FS utility runs
    util = str(out_unformat.get("utility", "")).lower()
    if util == "fs":
        if "FS_envy_avg" in out_unformat:
            for c in countries:
                out_form[c].loc["FS_envy_avg", years] = [out_unformat["FS_envy_avg"][(c, tt)] for tt in periods]
        if "FS_guilt_avg" in out_unformat:
            for c in countries:
                out_form[c].loc["FS_guilt_avg", years] = [out_unformat["FS_guilt_avg"][(c, tt)] for tt in periods]

    # Population + per-capita rows
    if L is not None:
        for c in countries:
            out_form[c].loc["L", years] = [L.at[c, tt] for tt in periods]
            out_form[c].loc["Y_pc", years] = [
                safe_div(out_unformat.get("Y", {}).get((c, tt), np.nan), L.at[c, tt], default=np.nan)
                for tt in periods
            ]
            out_form[c].loc["K_pc", years] = [
                safe_div(out_unformat.get("K", {}).get((c, tt), np.nan), L.at[c, tt], default=np.nan)
                for tt in periods
            ]
            out_form[c].loc["C_pc", years] = [
                safe_div(out_unformat.get("C", {}).get((c, tt), np.nan), L.at[c, tt], default=np.nan)
                for tt in periods
            ]
            if "U" in out_unformat:
                out_form[c].loc["U_pc", years] = [
                    safe_div(out_unformat["U"].get((c, tt), np.nan), L.at[c, tt], default=np.nan)
                    for tt in periods
                ]

    for c in countries:
        out_form[c] = _apply_spec(out_form[c], REGION_SPEC)
    out_form["global"] = _apply_spec(out_form["global"], GLOBAL_SPEC)

    return out_form

def results_to_excel(
    res: Dict[str, pd.DataFrame],
    countries: List[str],
    results_path: Path,
    filename: str,
    *,
    config: Dict[str, Any] | None = None,
) -> None:
    final_path = results_path / filename
    with pd.ExcelWriter(str(final_path), engine="xlsxwriter") as writer:
        if config is not None:
            # Sanitize config for display; avoid dumping large in-memory objects.
            cfg_pretty = _config_for_excel(config)
            config_str = yaml.dump(cfg_pretty, sort_keys=False)
            pd.DataFrame({"config": config_str.splitlines()}).to_excel(
                writer, sheet_name="config", index=False
            )
            worksheet = writer.sheets["config"]
            worksheet.set_column(0, 0, 60)

        for sheet in list(countries) + ["global"]:
            res[sheet].to_excel(writer, sheet_name=sheet)
            worksheet = writer.sheets[sheet]
            worksheet.set_column(0, 0, 28 if sheet != "global" else 40)

def plot_mu_for_run(
    mu_dict: Dict[Tuple[str, int], float],
    regions: List[str],
    periods: List[int],
    *,
    base_year: int,
    tstep: int,
    year_cap: int | None = None,
    region_names: Dict[str, str] | None = None,
    outfile: BytesIO | str | None = None,
) -> Tuple[int, int]:
    years = years_from_periods(periods, base_year, tstep=tstep)
    x_min = years[0]
    x_max = years[-1] if year_cap is None else min(years[-1], int(year_cap))

    cmap = plt.cm.get_cmap("tab20", len(regions))
    plt.figure(figsize=(8, 6))
    for idx, region in enumerate(regions):
        y = [mu_dict.get((region, t), np.nan) for t in periods]
        label = (region_names[region] if region_names else region)
        plt.plot(years, y, label=label, color=cmap(idx), linewidth=2)
    plt.xlim(x_min, x_max)
    plt.xlabel("Year")
    plt.ylabel("Abatement rate (μ)")
    plt.title("Abatement rate (μ) by region")
    plt.legend(bbox_to_anchor=(1.04, 1), loc="upper left", fontsize=10)
    plt.tight_layout(rect=[0, 0, 0.8, 1])
    if outfile is not None:
        plt.savefig(outfile, dpi=200, format="png")
    plt.close()
    return x_min, x_max

# ---------------------------------------------------------------------------
# Tabular run exports (planner/Nash/BAU)
# ---------------------------------------------------------------------------

def export_tabular(
    name: str,
    countries: List[str],
    solution: Dict[str, Any],
    periods: List[int],
    T: int,
    output_dir: Path,
    L: pd.DataFrame | None = None,
    *,
    tstep: int,
    base_year: int,
    backstop_switch_year: int | None = None,
    config: Dict[str, Any] | None = None,
) -> None:
    if not solution:
        logger.debug("No %s solution to export.", name)
        return
    if "disc" not in solution:
        raise ValueError(f"Run '{name}' missing 'disc' (discounted utilities required).")

    # Timestamped filename; optionally append a "disc" syllable for FS runs when FS discounting is enabled
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    util_key = str(solution.get("utility", "")).lower()
    disc_flag = False
    if util_key == "fs" and isinstance(config, dict):
        fs_enabled = bool(config.get("fs_disc_enabled", False))
        fs_mode = str(config.get("fs_disc_mode", "off")).strip().lower()
        # Add the "disc" syllable for FS runs when fs_disc_mode is not 'off'
        disc_flag = fs_enabled and (fs_mode != "off")
    name_out = f"{name}_disc" if disc_flag else name
    filename = f"{name_out}_{ts}.xlsx"

    res = output_format(countries, solution, periods, L, tstep=tstep, base_year=base_year)

    # Build a human-friendly config copy for the Excel 'config' sheet:
    #  - inject model, scenario, run_mode, and flavor explicitly,
    #  - keep their keys at the top for readability.
    cfg_for_excel: Dict[str, Any] = {}
    # Model: hard-coded for this codebase
    cfg_for_excel["model"] = "RICE13_FS"
    # Scenario: derive from run name, utility flavor, Negishi, and FS discount settings
    scenario_tag = _scenario_tag_for_run(
        run_name=name,
        util_key=util_key,
        solution=solution,
        config=config,
        disc_flag=disc_flag,
    )
    
    cfg_for_excel["scenario"] = scenario_tag

    # Derive run_mode from the run name (e.g. 'planner_crra' -> 'planner')
    run_mode = name.split("_", 1)[0] if name else "run"

    # Utility flavor: use util_key if provided, otherwise fall back to solution.
    # For BAU exports, use a neutral flavor label ('none').
    util = (util_key or str(solution.get("utility", ""))).lower().strip()
    if run_mode.startswith("bau"):
        flavor = "none"
    else:
        if util in {"crra", "fs"}:
            flavor = util
        else:
            flavor = "none"

    cfg_for_excel["run_mode"] = run_mode
    cfg_for_excel["flavor"] = flavor

    # Append the rest of the original config (if any), without overwriting
    # the core identity fields we just injected.
    for k, v in (config or {}).items():
        if k in ("model", "scenario", "run_mode", "flavor"):
            continue
        cfg_for_excel[k] = v

    results_to_excel(res, countries, output_dir, filename, config=cfg_for_excel)

    # μ plot
    buf = BytesIO()
    cap_year = None
    if backstop_switch_year is not None:
        cap_year = min(base_year + max(periods)*tstep, int(backstop_switch_year) + tstep)
    plot_mu_for_run(solution.get("mu", {}), countries, periods,
                    base_year=base_year, tstep=tstep, year_cap=cap_year,
                    region_names=None, outfile=buf)
    wb = load_workbook(output_dir / filename)
    ws = wb.create_sheet("mu Plot")
    buf.seek(0); ws.add_image(XLImage(PILImage.open(buf)), "A1")
    wb.save(output_dir / filename)

# ---------------------------------------------------------------------------
# Coalition helpers (Stability sheet)
# ---------------------------------------------------------------------------

# --- Flavor taxonomy & helpers (CRRA / FS-off / FS-file / FS-one_pass / FS-two_pass) ---

def _flavor_key(util: str, fs_variant: str | None) -> str:
    return util if util != "fs" or not fs_variant else f"fs_{fs_variant}"

def _flavors_to_export_from_config(config: Dict[str, Any] | None) -> List[Tuple[str, str | None]]:
    """
    Decide which flavors to export based on user's config (not on cache/list presence),
    so that two_pass runs won't export the intermediate one_pass artifacts.
    """
    out: List[Tuple[str, str | None]] = []
    cfg = config or {}
    if bool(cfg.get("run_coalition_crra", False)):
        out.append(("crra", None))
    if bool(cfg.get("run_coalition_fs", False)):
        mode = str(cfg.get("fs_disc_mode", "off")).strip().lower()
        if mode == "off":
            out.append(("fs", "off"))
        elif mode == "file":
            out.append(("fs", "file"))
        elif mode == "one_pass":
            out.append(("fs", "one_pass"))
        elif mode == "two_pass":
            out.append(("fs", "two_pass"))
        else:
            raise ValueError(f"Unknown fs_disc_mode={mode!r}")
    return out

def _full_neighbor_list(
    base_vec: List[int],
    countries: List[str],
    available_masks: set[str] | None = None,
) -> Tuple[List[List[int]], int, List[List[int]], List[List[int]]]:
    internal = list_internal_neighbors(base_vec)
    external = list_external_neighbors(base_vec)
    if available_masks is not None:
        to_mask = lambda v: "".join(str(int(x)) for x in v)
        internal = [v for v in internal if to_mask(v) in available_masks]
        external = [v for v in external if to_mask(v) in available_masks]
    full_list = internal + [base_vec] + external
    base_idx = len(internal)
    return full_list, base_idx, internal, external

def _write_stability_sheet(
    wb,
    *,
    members: str,
    full_list: List[List[int]],
    base_idx: int,
    internal_enabled: bool,
    external_enabled: bool,
    i_flags: List[int],
    e_flags: List[int],
    f_flags: List[int],
    i_wt: List[List[int]],
    leavers_idx: List[int],
    joiners_idx: List[int],
    i_missing_names: List[str],
    e_missing_names: List[str],
    countries: List[str],
) -> None:
    ws = wb.create_sheet("Stability")
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 20

    ws["B2"], ws["C2"] = "Coalition Members", members

    if not internal_enabled:
        ws["B4"], ws["C4"] = "Internally Stable", "unchecked"
    else:
        ws["B4"], ws["C4"] = "Internally Stable", bool(i_flags[base_idx])

    ws["D4"], ws["E4"] = "Internal Deviators (would leave)", (
        ",".join(countries[i] for i in leavers_idx) if leavers_idx else ""
    )
    # Explicit list of missing internal neighbors (warning-only)
    ws["B5"], ws["C5"] = "Internal checks missing", (
        ",".join(i_missing_names) if i_missing_names else ""
    )
    if i_missing_names:
        # Highlight missing internal checks in red so users see that
        # "True" above should be interpreted as "maybe" only.
        ws["C5"].font = Font(color="FF0000")

    if not external_enabled:
        ws["B6"], ws["C6"] = "Externally Stable", "unchecked"
    else:
        ws["B6"], ws["C6"] = "Externally Stable", bool(e_flags[base_idx])
    ws["D6"], ws["E6"] = "External Deviators (would join)", (
        ",".join(countries[i] for i in joiners_idx) if joiners_idx else ""
    )
    # Explicit list of missing external neighbors (warning-only)
    ws["B7"], ws["C7"] = "External checks missing", (
        ",".join(e_missing_names) if e_missing_names else ""
    )
    if e_missing_names:
        # Highlight missing external checks in red (same semantics as internal).
        ws["C7"].font = Font(color="FF0000")

    if not internal_enabled and not external_enabled:
        ws["B8"], ws["C8"] = "Fully Stable", "unchecked"
    else:
        ws["B8"], ws["C8"] = "Fully Stable", bool(f_flags[base_idx])

    # Explanatory note to make the "maybe" interpretation explicit on the sheet.
    ws["B10"] = (
        "Note: Stability booleans above consider only coalitions that were actually solved. "
        "If any region names appear in the 'checks missing' rows, interpret True as 'maybe stable'."
    )
    ws["B10"].font = Font(italic=True)


def _write_stability_deltas_sheet(
    wb,
    *,
    base_vec: List[int],
    countries: List[str],
    payoff_by_mask: Dict[str, List[float]],
    eps: float,
) -> None:
    """
    Write a compact, audit-friendly table that shows *exactly* the per-region
    payoff objects used in the SSOT stability test:

      payoff_r(S) = PV_r = sum_t disc[r,t] * U[r,t]

    For each one-bit neighbor coalition (leave/join), we report only the deviator's
    base PV, neighbor PV, and Δ=neighbor-base, plus Missing?/Deviates? flags.
    """
    ws = wb.create_sheet("Stability deltas")
    ws.freeze_panes = "A3"

    # Column widths for readability
    ws.column_dimensions["A"].width = 10  # Move
    ws.column_dimensions["B"].width = 10  # Region
    ws.column_dimensions["C"].width = 16  # Neighbor mask
    ws.column_dimensions["D"].width = 16  # Base PV
    ws.column_dimensions["E"].width = 16  # Neighbor PV
    ws.column_dimensions["F"].width = 16  # Delta
    ws.column_dimensions["G"].width = 10  # eps
    ws.column_dimensions["H"].width = 12  # Deviates?
    ws.column_dimensions["I"].width = 10  # Missing?

    # Normalize base_vec defensively
    n = len(countries)
    bvec = [1 if int(x) else 0 for x in base_vec]
    if len(bvec) != n:
        raise ValueError(f"base_vec length {len(bvec)} != len(countries) {n}")

    base_mask = _mask_of(bvec)
    base_pay = payoff_by_mask.get(base_mask)
    if base_pay is None:
        raise KeyError(f"Base coalition {base_mask} missing from payoff map.")

    # Header block
    ws["A1"], ws["B1"] = "Base mask", base_mask
    ws["D1"], ws["E1"] = "eps", float(eps)
    ws["A2"] = "Move"
    ws["B2"] = "Region"
    ws["C2"] = "Neighbor mask"
    ws["D2"] = "Base PV"
    ws["E2"] = "Neighbor PV"
    ws["F2"] = "Δ (neighbor-base)"
    ws["G2"] = "eps"
    ws["H2"] = "Deviates?"
    ws["I2"] = "Missing?"
    for cell in ("A2","B2","C2","D2","E2","F2","G2","H2","I2"):
        ws[cell].font = Font(bold=True)

    row = 3

    def _write_row(move: str, i: int, neigh_mask: str, neigh_pay: List[float] | None):
        nonlocal row
        base_i = float(base_pay[i])
        missing = neigh_pay is None
        if missing:
            neigh_i = np.nan
            delta = np.nan
            deviates = False
        else:
            neigh_i = float(neigh_pay[i])
            delta = neigh_i - base_i
            deviates = bool(delta > float(eps))

        ws.cell(row=row, column=1, value=move)               # A
        ws.cell(row=row, column=2, value=str(countries[i]))  # B
        ws.cell(row=row, column=3, value=neigh_mask)         # C
        ws.cell(row=row, column=4, value=base_i)             # D
        ws.cell(row=row, column=5, value=neigh_i)            # E
        ws.cell(row=row, column=6, value=delta)              # F
        ws.cell(row=row, column=7, value=float(eps))         # G
        ws.cell(row=row, column=8, value=deviates)           # H
        ws.cell(row=row, column=9, value=missing)            # I

        # Light formatting: highlight actual profitable deviations
        if deviates:
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = Font(color="FF0000", bold=True)
        # Grey out missing neighbor rows (audit cue)
        if missing:
            ws.cell(row=row, column=9).font = Font(color="808080", italic=True)
            ws.cell(row=row, column=5).font = Font(color="808080", italic=True)
            ws.cell(row=row, column=6).font = Font(color="808080", italic=True)

        row += 1

    # Internal neighbors: each member i leaves
    for i, bit in enumerate(bvec):
        if bit != 1:
            continue
        neigh = bvec.copy()
        neigh[i] = 0
        nm = _mask_of(neigh)
        _write_row("leave", i, nm, payoff_by_mask.get(nm))

    # External neighbors: each outsider i joins
    for i, bit in enumerate(bvec):
        if bit != 0:
            continue
        join = bvec.copy()
        join[i] = 1
        jm = _mask_of(join)
        _write_row("join", i, jm, payoff_by_mask.get(jm))


# ---------------------------------------------------------------------------
# Cache-first coalition writer (uses SSOT stability, legacy layout)
# ---------------------------------------------------------------------------

def write_from_cache(
    store: CoalitionStore,
    *,
    countries: List[str],
    periods: List[int],
    output_dir: Path,
    utility: str = "crra",
    L: pd.DataFrame | None = None,
    tstep: int = 10,
    base_year: int = 2015,
    config: Dict[str, Any] | None = None,
    selection: Iterable[Tuple[int, ...]] | None = None,
    backstop_switch_year: int | None = None,
    export_selection: Iterable[Tuple[int, ...]] | None = None,
) -> Dict[str, List[float]]:
    output_dir.mkdir(parents=True, exist_ok=True)

    # Build the spec_id that matches cache rows we want to export (utility- and discount-aware)
    T = int(max(periods))
    neg_use = bool((config or {}).get("negishi_use", False))

    # Strict per-utility coalition S-policy; do not read legacy/shared keys.
    exoS_df = None

    if utility.lower() == "crra":
        from RICE13_FS.common.utils import resolve_coalition_S_for_export
        # Determine coalition-CRRA S mode
        s_mode = str((config or {}).get("coalition_crra_S_mode", "optimal")).strip().lower()

        # Provide minimal params if BAU is requested (expects .bau_saving_rates)
        params_ns = None
        if s_mode == "bau":
            params_ns = SimpleNamespace(bau_saving_rates=(config or {}).get("bau_saving_rates"))

        # Pre-load exogenous S if FILE mode is requested (helper avoids I/O by design)
        exoS_pre = None
        if s_mode == "file":
            s_path = (config or {}).get("coalition_crra_S_file")
            if not s_path:
                raise ValueError("coalition_crra_S_mode='file' requires coalition_crra_S_file in config")
            s_df = _read_exog_S_csv(Path(s_path))
            exoS_pre = normalize_exogenous_S(s_df, countries, T)

        # Planner FS needs the exact planner discount tag
        planner_disc_tag = None
        if s_mode == "planner_fs":
            planner_disc_tag = (config or {}).get("planner_fs_disc_tag")
            if not planner_disc_tag:
                # reasonable fallback if you tagged the planner with the plain fs_disc_mode
                mode_hint = str((config or {}).get("fs_disc_mode", "")).strip().lower()
                if mode_hint in {"one_pass", "two_pass", "file"}:
                    planner_disc_tag = mode_hint

        # Resolve S exactly like the solver (keyword-only helper)
        exoS_df = resolve_coalition_S_for_export(
            s_mode=s_mode,
            params=params_ns,
            countries=countries,
            T=T,
            store=store,
            negishi_use=neg_use,
            negishi_weights=(config or {}).get("negishi_weights"),
            population_weight_envy_guilt=bool((config or {}).get("population_weight_envy_guilt", False)),
            planner_disc_tag=planner_disc_tag,
            exoS_df=exoS_pre,
        )

    elif utility.lower() == "fs":
        # FS export does not normally need exoS_df, but keep minimal handling for completeness.
        s_mode = str((config or {}).get("coalition_fs_S_mode", "")).strip().lower()
        if s_mode == "file":
            s_path = (config or {})["coalition_fs_S_file"]
            s_df = _read_exog_S_csv(Path(s_path))
            exoS_df = normalize_exogenous_S(s_df, countries, T)
        elif s_mode == "bau":
            # If you injected BAU S into config earlier, use it; otherwise leave None.
            exoS_df = (config or {}).get("bau_saving_rates")
        # else: optimal/planner_* not required for FS export keying; rows carry spec_id explicitly.
    else:
        raise ValueError(f"Unknown coalition utility in cache export: {utility!r}")

    # --- Row selection strategy ---
    # CRRA: keep strict spec_id lookup.
    # FS  : scan whole store, but DO NOT rely on r.meta (workers don't always write utility/disc_tag there).
    if utility.lower() == "fs":
        mode = str((config or {}).get("fs_disc_mode", "off")).strip().lower()
        want_aligned = mode in {"one_pass", "two_pass"}
        want_file = (mode == "file")
        logger.debug("[EXPORT][FS] scan start | mode=%s (aligned=%s file=%s)", mode, want_aligned, want_file)

        # Candidate rows: same dimensionality; optionally match selection
        cand = [
            r for r in store.iter_rows(spec_id=None)
            if (r.vector and len(r.vector) == len(countries) and any(int(x) for x in r.vector))
        ]
        logger.debug("[EXPORT][FS] candidates by dim: %d", len(cand))

        if selection is not None:
            sel_masks = {"".join(str(int(x)) for x in vec) for vec in selection}
            cand = [r for r in cand if r.vector and "".join(str(int(x)) for x in r.vector) in sel_masks]

        seen_total = seen_util = seen_tagmatch = 0
        rows = []

        # Vet candidates by reading the full solution and (if missing) falling back to metadata
        for r in cand:
            seen_total += 1
            vec_t = tuple(r.vector) if r.vector is not None else None
            sid = getattr(r, "spec_id", None)
            rec = store.get(vec_t, sid) if sid is not None else {}
            sol  = (rec or {}).get("solution") or {}
            meta = (rec or {}).get("meta") or {}

            # NEW: fallback to meta when blob is missing
            util_sol = str(sol.get("utility") or meta.get("utility", "")).lower()
            if util_sol != "fs":
                logger.debug("[EXPORT][FS] skip vec=%s spec=%s: utility=%s",
                             "".join(map(str, r.vector or [])), sid, util_sol)
                continue
            seen_util += 1
            tag = str(sol.get("disc_tag") or meta.get("disc_tag", "")).lower()
            # --- Strict tag enforcement & early filtering ---
            # We require the disc_tag to match the requested fs_disc_mode exactly,
            # rather than admitting any "aligned" FS tag.
            if want_aligned:
                # Exact mode match: only "one_pass" OR only "two_pass"
                mode_val = mode  # from above: str((config or {}).get("fs_disc_mode", "off")).lower()
                if mode_val == "one_pass":
                    tag_ok = ("one_pass" in tag)
                elif mode_val == "two_pass":
                    tag_ok = ("two_pass" in tag)
                else:
                    # Defensive: treat unknown as not ok
                    tag_ok = False
                if not tag_ok:
                    logger.debug("[EXPORT][FS] skip vec=%s spec=%s: tag=%s !exact(%s)",
                                 "".join(map(str, r.vector or [])), sid, tag, mode_val)
                    continue
            elif want_file:
                # File-tagged FS runs must carry a disc:file:* tag
                if not tag.startswith("disc:file:"):
                    logger.debug("[EXPORT][FS] skip vec=%s spec=%s: tag=%s !file",
                                 "".join(map(str, r.vector or [])), sid, tag)
                    continue
            else:
                # off/data → allow empty or 'data' tags only
                if tag and ("data" not in tag):
                    logger.debug("[EXPORT][FS] skip vec=%s spec=%s: tag=%s !off/data",
                                 "".join(map(str, r.vector or [])), sid, tag)
                    continue

            # Passed strict utility + tag checks → accept
            seen_tagmatch += 1
            rows.append(r)

        logger.info("[EXPORT][FS] admitted rows: %d (candidates=%d util_ok=%d tag_ok=%d)",
                    len(rows), seen_total, seen_util, seen_tagmatch)

    else:
        # CRRA → build one spec_id and stream rows from that id.
        spec_id = build_solution_spec_id(
            utility="crra",
            T=T,
            countries=countries,
            population_weight_envy_guilt=False,
            exogenous_S=exoS_df,
            negishi_use=neg_use,
            negishi_weights=(config or {}).get("negishi_weights"),
            disc_tag=None,
        )
        rows = [
            r for r in store.iter_rows(spec_id=spec_id)
            if (r.vector and any(int(x) for x in r.vector))
        ]
        logger.info("[EXPORT][CRRA] rows under spec_id=%s: %d", spec_id, len(rows))
        if selection is not None:
            sel_masks = {"".join(str(int(x)) for x in vec) for vec in selection}
            rows = [r for r in rows if r.vector and "".join(str(int(x)) for x in r.vector) in sel_masks]

    # Build payoff map ONLY from solutions we actually accept (esp. FS filtering above).
    payoff_by_mask: Dict[str, List[float]] = {}
    export_masks: set[str] | None = None
    if export_selection is not None:
        export_masks = {"".join(str(int(x)) for x in vec) for vec in export_selection}

    for r in rows:
        vec = list(r.vector)
        vec_mask = "".join(str(int(v)) for v in vec)
        members = coalition_vec_to_member_string(vec, countries)

        # Resolve the cached record/solution for this row
        if utility.lower() == "fs":
            sid = getattr(r, "spec_id", None)
            rec = store.get(tuple(vec), sid) if sid is not None else {}
        else:
            rec = store.get(tuple(vec), spec_id) or {}

        solution = rec.get("solution") or {}
        if "disc" not in solution:
            raise ValueError("Cached solution missing 'disc'; cache must store discounted-ready solutions.")

        # --- DIAGNOSTICS (fixed, no stray params) ---
        logger.debug(
            "[EXPORT] using vec=%s | util=%s | tag=%s | spec=%s",
            "".join(map(str, vec)),
            str(solution.get("utility", "")).lower(),
            str(solution.get("disc_tag", "")).lower(),
            getattr(r, "spec_id", None),
        )

        # Utility and suffix from the actual solution (not from config/param), to avoid overwrites.
        util_sol = str(solution.get("utility", utility)).lower()
        fs_suffix = ""
        if util_sol == "fs":
            tag = str(solution.get("disc_tag", "")).lower()
            if "one_pass" in tag:
                fs_suffix = "_one_pass"
            elif "two_pass" in tag:
                fs_suffix = "_two_pass"
            else:
                # If config claims aligned FS but we cannot infer, error
                if isinstance(config, dict) and bool(config.get("fs_disc_enabled", False)) \
                   and str(config.get("fs_disc_mode", "off")).strip().lower() in {"one_pass", "two_pass"}:
                    raise ValueError("Cache export: FS coalition appears aligned but disc_tag lacks one_pass/two_pass.")

        # Record payoff row now that we know this coalition belongs to this flavor
        payoff_by_mask[vec_mask] = r.payoff

        # If this coalition is not requested for export, skip file writing
        if export_masks is not None and vec_mask not in export_masks:
            continue

        ts = datetime.now().strftime("%Y%m%d_%H%M")
        neg_suffix = "_N" if (bool((config or {}).get("negishi_use", False)) or ("negishi_weights" in solution)) else ""
        filename = f"coal_{util_sol}{fs_suffix}{neg_suffix}_{members}_{ts}.xlsx"
        res = output_format(countries, solution, periods, L=L, tstep=tstep, base_year=base_year)
        results_to_excel(res, countries, output_dir, filename, config=config)

        wb = load_workbook(output_dir / filename)
        _attach_mu_plot(
            wb, solution, countries, periods,
            base_year=base_year, tstep=tstep, backstop_switch_year=backstop_switch_year
        )

        # Stability (SSOT)
        eps = float((config or {}).get("stability_eps", 1e-3))
        from_mask = {"".join(str(int(v)) for v in rr.vector): rr.payoff for rr in rows}
        stab = evaluate_stability_for_mask(vec, countries=countries, payoff_by_mask=from_mask, eps=eps)

        full_list, base_idx, internal, external = _full_neighbor_list(
            vec, countries, available_masks=set(from_mask.keys())
        )
        i_flags = [0] * len(full_list)
        e_flags = [0] * len(full_list)
        f_flags = [0] * len(full_list)
        i_flags[base_idx] = 1 if stab.internally_stable else 0
        e_flags[base_idx] = 1 if stab.externally_stable else 0
        f_flags[base_idx] = 1 if stab.fully_stable else 0

        _write_stability_sheet(
            wb,
            members=members,
            full_list=full_list,
            base_idx=base_idx,
            internal_enabled=bool(config and config.get("coalition_check_internal", False)),
            external_enabled=bool(config and config.get("coalition_check_external", False)),
            i_flags=i_flags, e_flags=e_flags, f_flags=f_flags, i_wt=[],
            leavers_idx=list(stab.leavers_idx), joiners_idx=list(stab.joiners_idx),
            i_missing_names=list(stab.i_missing_names), e_missing_names=list(stab.e_missing_names),
            countries=countries,
        )
        
        # New: compact audit table showing the exact deviator PV utilities and deltas
        # used by the SSOT stability test.
        _write_stability_deltas_sheet(
            wb,
            base_vec=vec,
            countries=countries,
            payoff_by_mask=from_mask,
            eps=eps,
        )

        wb.save(output_dir / filename)
        logger.info("Wrote %s (cache-first).", output_dir / filename)

    return payoff_by_mask


# ---------------------------------------------------------------------------
# Stability overview from payoff map (streaming path, per flavor)
# ---------------------------------------------------------------------------

def export_stability_overview_from_map(
    util_key: str,
    payoff_by_mask: Dict[str, List[float]],
    *,
    countries: List[str],
    periods: List[int],
    output_dir: Path,
    config: Dict[str, Any] | None = None,
) -> None:
    """Write one stability overview XLSX for a single flavor using a small payoff map.

    This uses evaluate_stability_for_mask(...) as the SSOT and distinguishes between
    "certain" vs "maybe" stability:
      - Internally/externally stable (certain): no deviators found AND no missing neighbors
        in that direction.
      - Internally/externally stable (maybe): no deviators found, but at least one
        neighbor coalition in that direction is missing from the payoff map.
      - Fully stable (certain): internally & externally stable AND no missing neighbors
        in either direction.
      - Fully stable (maybe): fully stable but relying on at least one missing neighbor.
    """
    if not payoff_by_mask:
        logger.info("Skipping overview (%s): empty payoff map.", util_key)
        return

    negishi = bool(config.get("negishi_use")) if isinstance(config, dict) else False
    eps = float((config or {}).get("stability_eps", 1e-3))

    # Coverage diagnostics (non-empty coalitions only)
    n = len(countries)
    present = set(payoff_by_mask.keys())
    present_nonempty = {m for m in present if "1" in m}
    expected_nonempty = (1 << n) - 1
    coverage = (len(present_nonempty) / expected_nonempty) if expected_nonempty else 1.0
    logger.info(
        "Overview (%s): %d/%d non-empty coalitions present (coverage=%.3f).",
        util_key, len(present_nonempty), expected_nonempty, coverage,
    )

    # Enumerate all possible non-empty coalitions for this region set and determine missing ones.
    all_masks = [format(k, f"0{n}b") for k in range(1, 1 << n)]
    missing_masks = [m for m in all_masks if m not in present_nonempty]

    # Optional: read failures.txt to attach a coarse failure reason per missing coalition (best-effort only).
    reason_by_mask: Dict[str, str] = {}
    try:
        results_dir = Path((config or {}).get("results_dir", output_dir))
        failures_path = results_dir / "failures.txt"
        if failures_path.exists():
            with open(failures_path, "r", encoding="utf-8") as fh:
                for line in fh:
                    parts = line.rstrip("\\n").split("\\t")
                    if len(parts) < 3:
                        continue
                    mask_field = parts[2].strip()
                    # Heuristic: treat a pure 0/1 string of correct length as a coalition mask.
                    if len(mask_field) == n and set(mask_field) <= {"0", "1"}:
                        reason = parts[-1].strip() if len(parts) >= 4 else ""
                        # Last occurrence wins; that is fine for diagnostics.
                        reason_by_mask[mask_field] = reason
    except Exception:
        logger.debug("Could not read failures.txt for overview diagnostics.", exc_info=True)

    # Build rows by iterating masks in the payoff map (non-empty coalitions only)
    coal_rows: List[Dict[str, Any]] = []
    # Deterministic ordering: by coalition size then lexicographic mask
    masks_sorted = sorted(list(present_nonempty), key=lambda s: (s.count("1"), s))
    for base_mask in masks_sorted:
        base_vec = [1 if ch == "1" else 0 for ch in base_mask]
        members = coalition_vec_to_member_string(base_vec, countries)
        res = evaluate_stability_for_mask(
            base_vec, countries=countries, payoff_by_mask=payoff_by_mask, eps=eps
        )

        # Derive "certain" vs "maybe" flags from missing-neighbor information.
        internal_certain = res.internally_stable and not res.i_missing_names
        internal_maybe = res.internally_stable and bool(res.i_missing_names)

        external_certain = res.externally_stable and not res.e_missing_names
        external_maybe = res.externally_stable and bool(res.e_missing_names)

        full_certain = res.fully_stable and internal_certain and external_certain
        full_maybe = res.fully_stable and not full_certain

        if logger.isEnabledFor(logging.DEBUG):
            logger.debug(
                "[STAB][%s] %s members=%s | int=%s ext=%s | leavers=%s | joiners=%s | "
                "int_missing=%s | ext_missing=%s",
                util_key,
                base_mask,
                members,
                res.internally_stable,
                res.externally_stable,
                ",".join(countries[i] for i in res.leavers_idx) or "–",
                ",".join(countries[i] for i in res.joiners_idx) or "–",
                ",".join(res.i_missing_names) or "–",
                ",".join(res.e_missing_names) or "–",
            )
            _debug_dump_internal(base_vec, countries, payoff_by_mask, eps=eps)
            _debug_dump_external(base_vec, countries, payoff_by_mask, eps=eps)

        coal_rows.append(
            {
                "mask": base_mask,
                "name": members,
                "size": int(sum(base_vec)),
                # SSOT stability flags (based only on available neighbors)
                "InternallyStable": res.internally_stable,
                "ExternallyStable": res.externally_stable,
                "FullyStable": res.fully_stable,
                "Leavers": ",".join(countries[i] for i in res.leavers_idx),
                "Joiners": ",".join(countries[i] for i in res.joiners_idx),
                # New: diagnostics on missing neighbors
                "InternalMissing": ",".join(res.i_missing_names),
                "ExternalMissing": ",".join(res.e_missing_names),
                # New: certain vs maybe classification
                "InternallyStableCertain": internal_certain,
                "InternallyStableMaybe": internal_maybe,
                "ExternallyStableCertain": external_certain,
                "ExternallyStableMaybe": external_maybe,
                "FullyStableCertain": full_certain,
                "FullyStableMaybe": full_maybe,
            }
        )

    df = pd.DataFrame(coal_rows)
    if not df.empty:
        df["mask_int"] = df["mask"].apply(lambda b: int(str(b), 2))
        df = (
            df.sort_values(["size", "mask_int"])
            .drop(columns=["mask_int"])
            .reset_index(drop=True)
        )

    # Top-level summary (counts and shares, certain vs maybe)
    total_present = len(df)
    nonempty_possible = expected_nonempty
    summary_rows: List[Dict[str, Any]] = []

    summary_rows.append(
        {
            "Metric": "Non-empty coalitions present",
            "Count": total_present,
            "Share": 1.0,
        }
    )
    summary_rows.append(
        {
            "Metric": "Non-empty coalitions possible",
            "Count": nonempty_possible,
            "Share": 1.0,
        }
    )
    summary_rows.append(
        {
            "Metric": "Coverage (present / possible)",
            "Count": total_present,
            "Share": coverage,
        }
    )

    if total_present:
        s_int_certain = int(df["InternallyStableCertain"].sum())
        s_int_maybe = int(df["InternallyStableMaybe"].sum())
        s_ext_certain = int(df["ExternallyStableCertain"].sum())
        s_ext_maybe = int(df["ExternallyStableMaybe"].sum())
        s_full_certain = int(df["FullyStableCertain"].sum())
        s_full_maybe = int(df["FullyStableMaybe"].sum())

        summary_rows.extend(
            [
                {
                    "Metric": "Internally stable (certain)",
                    "Count": s_int_certain,
                    "Share": s_int_certain / total_present,
                },
                {
                    "Metric": "Internally stable (maybe only)",
                    "Count": s_int_maybe,
                    "Share": s_int_maybe / total_present,
                },
                {
                    "Metric": "Externally stable (certain)",
                    "Count": s_ext_certain,
                    "Share": s_ext_certain / total_present,
                },
                {
                    "Metric": "Externally stable (maybe only)",
                    "Count": s_ext_maybe,
                    "Share": s_ext_maybe / total_present,
                },
                {
                    "Metric": "Fully stable (certain)",
                    "Count": s_full_certain,
                    "Share": s_full_certain / total_present,
                },
                {
                    "Metric": "Fully stable (maybe only)",
                    "Count": s_full_maybe,
                    "Share": s_full_maybe / total_present,
                },
            ]
        )

    summary = pd.DataFrame(summary_rows)

    # Summary by size: one compact table with totals and certain/maybe counts
    if total_present:
        grp = df.groupby("size")
        by_size = grp["mask"].count().to_frame(name="CoalitionsPresent")
        by_size["InternallyStableCertain"] = grp["InternallyStableCertain"].sum()
        by_size["InternallyStableMaybe"] = grp["InternallyStableMaybe"].sum()
        by_size["ExternallyStableCertain"] = grp["ExternallyStableCertain"].sum()
        by_size["ExternallyStableMaybe"] = grp["ExternallyStableMaybe"].sum()
        by_size["FullyStableCertain"] = grp["FullyStableCertain"].sum()
        by_size["FullyStableMaybe"] = grp["FullyStableMaybe"].sum()
        by_size = by_size.reset_index()
    else:
        by_size = pd.DataFrame(
            columns=[
                "size",
                "CoalitionsPresent",
                "InternallyStableCertain",
                "InternallyStableMaybe",
                "ExternallyStableCertain",
                "ExternallyStableMaybe",
                "FullyStableCertain",
                "FullyStableMaybe",
            ]
        )

    # Long "By size" sheet (uses SSOT stability flags; internal split into certain/maybe)
    def _names_for(df0: pd.DataFrame, colflag: str, name_prefix: str) -> pd.DataFrame:
        """
        Build a wide table of coalition names by size for a given boolean flag
        column (e.g. 'InternallyStableCertain').

        Each set of name columns gets its own prefix so that subsequent merges
        on 'size' do not hit pandas' duplicate-column/suffix guard.
        """
        sub = df0.loc[df0[colflag], ["size", "name"]]
        names_by_size = sub.groupby("size")["name"].apply(list)
        out_rows: List[Dict[str, Any]] = []
        for k in sorted(df0["size"].unique()):
            L = names_by_size.get(k, [])
            row: Dict[str, Any] = {"size": k, "count": len(L)}
            for j, nm in enumerate(L, 1):
                row[f"{name_prefix}_{j}"] = nm
            out_rows.append(row)
        return pd.DataFrame(out_rows)

    # Total coalitions by size
    by_size_all = (
        df.groupby("size")["mask"]
        .count()
        .reset_index()
        .rename(columns={"mask": "Coalitions"})
    )
    # Internally stable: split into certain vs maybe-only
    by_size_int_certain = _names_for(df, "InternallyStableCertain", "int_certain").rename(
        columns={"count": "InternallyStableCertain"}
    )
    by_size_int_maybe = _names_for(df, "InternallyStableMaybe", "int_maybe").rename(
        columns={"count": "InternallyStableMaybe"}
    )
    # Externally & fully stable
    by_size_ext  = _names_for(df, "ExternallyStable", "ext").rename(
        columns={"count": "ExternallyStable"}
    )
    by_size_full = _names_for(df, "FullyStable", "full").rename(
        columns={"count": "FullyStable"}
    )

    def _merge_on_size(frames):
        return reduce(lambda a,b: a.merge(b, on="size", how="left"), frames)

    by_size_sheet = _merge_on_size(
        [by_size_all, by_size_int_certain, by_size_int_maybe, by_size_ext, by_size_full]
    ).fillna("")

    # Region sheet — fully stable coalitions per region (from masks directly)
    region_rows = []
    full_df = df.loc[df["FullyStable"], ["mask","name","size"]]
    for ridx, r in enumerate(countries):
        names = []
        for _, row in full_df.iterrows():
            mask = row["mask"]
            if mask[ridx] == "1":
                names.append(row["name"])
        out = {"Region": r, "FullyStableMemberships": len(names)}
        for j, nm in enumerate(names, 1):
            out[f"name_{j}"] = nm
        region_rows.append(out)
    by_region_sheet = pd.DataFrame(region_rows).fillna("")

    # Missing coalitions table (diagnostic; one line per coalition)
    missing_rows: List[Dict[str, Any]] = []
    for mask in missing_masks:
        vec = [1 if ch == "1" else 0 for ch in mask]
        members = coalition_vec_to_member_string(vec, countries)
        missing_rows.append(
            {
                "mask": mask,
                "name": members,
                "size": int(sum(vec)),
                "FailureReason": reason_by_mask.get(mask, ""),
            }
        )
    missing_df = (
        pd.DataFrame(missing_rows)
        .sort_values(["size", "mask"])
        .reset_index(drop=True)
        if missing_rows
        else pd.DataFrame(columns=["mask", "name", "size", "FailureReason"])
    )


    ts = datetime.now().strftime("%Y%m%d_%H%M")
    label = f"stability_overview_{util_key}{'_N' if negishi else ''}_{ts}.xlsx"
    outpath = output_dir / label
    with pd.ExcelWriter(outpath, engine="openpyxl") as xw:
        # Summary sheet: overview at top, missing coalitions appended below
        summary.to_excel(xw, sheet_name="Summary", index=False)
        if not missing_df.empty:
            start_row = len(summary) + 2  # one blank row between tables
            missing_df.to_excel(
                xw, sheet_name="Summary", index=False, startrow=start_row
            )

        # Compact per-size summary
        by_size.to_excel(xw, sheet_name="Summary_by_size", index=False)
        # Detailed coalition list and names-by-size/region sheets
        df.to_excel(xw, sheet_name="Coalitions", index=False)
        by_size_sheet.to_excel(xw, sheet_name="By size", index=False)
        by_region_sheet.to_excel(xw, sheet_name="Region", index=False)

        ws_summary = xw.sheets["Summary"]
        ws_summary.column_dimensions["A"].width = 28
        ws_sbs = xw.sheets["Summary_by_size"]
        for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
            ws_sbs.column_dimensions[col].width = 20
        ws_coal = xw.sheets["Coalitions"]
        ws_coal.column_dimensions["A"].width = 15
        ws_coal.column_dimensions["B"].width = 15
        ws_coal.column_dimensions["D"].width = 15
        ws_coal.column_dimensions["E"].width = 15
        ws_bsz = xw.sheets["By size"]
        ws_bsz.column_dimensions["C"].width = 15
        ws_bsz.column_dimensions["G"].width = 15

        ws_reg = xw.sheets["Region"]
        ws_reg.column_dimensions["B"].width = 27

    logger.info("Wrote stability overview (%s) to %s", util_key, outpath)


# ---------------------------------------------------------------------------
# Top-level orchestrator matching CLI (unchanged call shape)
# ---------------------------------------------------------------------------

def export_all(
    *,
    coop_solution: Dict[str, Any] | None,
    noncoop_solution: Dict[str, Any] | None,
    coalitions: List[Dict[str, Any]] | None,
    countries: List[str],
    periods: List[int],
    T: int,
    output_dir: Path,
    L: pd.DataFrame | None,
    tstep: int,
    base_year: int,
    backstop_switch_year: int | None = None,
    bau_solution: Dict[str, Any] | None = None,
    config: Dict[str, Any] | None = None,
) -> None:
    def _planner_name(util: str, sol: dict | None) -> str:
        return f"planner{'_N' if (sol and ('negishi_weights' in sol)) else ''}_{util}"
    def _nash_name(util: str) -> str:
        return f"nash_{util}"

    # BAU
    if bau_solution:
        if "disc" not in bau_solution:
            raise ValueError("BAU solution missing 'disc'; discounted utilities are required.")
        export_tabular("bau", countries, bau_solution, periods, T=max(periods),
                       output_dir=output_dir, L=L, tstep=tstep, base_year=base_year,
                       backstop_switch_year=backstop_switch_year, config=config)

    # Planners
    if coop_solution:
        for util_key, sol in (coop_solution or {}).items():
            if sol is None:
                continue
            if "disc" not in sol:
                raise ValueError(f"Planner '{util_key}' missing 'disc'.")
            export_tabular(_planner_name(util_key, sol), countries, sol, periods, T=max(periods),
                           output_dir=output_dir, L=L, tstep=tstep, base_year=base_year,
                           backstop_switch_year=backstop_switch_year, config=config)

    # Nash
    if noncoop_solution:
        for util_key, sol in (noncoop_solution or {}).items():
            if sol is None:
                continue
            if "disc" not in sol:
                raise ValueError(f"Nash '{util_key}' missing 'disc'.")
            export_tabular(_nash_name(util_key), countries, sol, periods, T=max(periods),
                           output_dir=output_dir, L=L, tstep=tstep, base_year=base_year,
                           backstop_switch_year=backstop_switch_year, config=config)

    # --- STREAMING PATH: export coalitions by flavor from the cache ---
    coalition_spec = None
    if isinstance(config, dict):
        coalition_spec = str(config.get("coalition", "none"))
    _spec = (coalition_spec or "").strip().lower() if isinstance(coalition_spec, str) else ""

    # If no coalition export requested, we’re done.
    if not (_spec and _spec != "none"):
        return

    # Determine flavors to export from config (not from cache presence).
    flavors = _flavors_to_export_from_config(config)
    if not flavors:
        logger.info("No coalition flavors selected in config; skipping coalition export.")
        return

    # Open the coalition store for the current run fingerprint
    cfg = config or {}
    cache_dir = Path(cfg["cache_dir"])
    cache_ns = str(cfg["cache_namespace"])
    fingerprint = build_config_fingerprint(cfg, countries)
    store = CoalitionStore(cache_dir, cache_ns, fingerprint)

    # Determine selection (all masks vs a specific coalition)
    if _spec in {"all", "*"}:
        # Mega run: stream all rows for each flavor/spec_id, and export all.
        selection = None
        export_selection = None
    else:
        # Single-coalition run:
        # - selection: base coalition + all 1-bit neighbors (internal + external)
        #   used to build the payoff map for stability.
        # - export_selection: base coalition only, used to limit which
        #   coalition workbooks are written.
        base_vec = parse_coalition_spec(coalition_spec, regions=countries)
        internal_neigh = list_internal_neighbors(base_vec)
        external_neigh = list_external_neighbors(base_vec)

        seen: set[Tuple[int, ...]] = set()
        selection_list: List[Tuple[int, ...]] = []
        for v in [base_vec] + internal_neigh + external_neigh:
            t = tuple(int(x) for x in v)
            if t not in seen:
                seen.add(t)
                selection_list.append(t)
        selection = selection_list
        export_selection = [tuple(base_vec)]

    # Export per flavor
    for util, fs_variant in flavors:
        # Prepare a flavor-specific copy of config to drive spec_id and suffix checks
        cfg_flavor = dict(cfg)
        if util == "fs":
            if fs_variant in {"one_pass", "two_pass"}:
                cfg_flavor["fs_disc_enabled"] = True
                cfg_flavor["fs_disc_mode"] = fs_variant
            elif fs_variant == "file":
                cfg_flavor["fs_disc_enabled"] = True
                cfg_flavor["fs_disc_mode"] = "file"
            else:  # "off"
                cfg_flavor["fs_disc_enabled"] = False
                cfg_flavor["fs_disc_mode"] = "off"

        # Flavor banner
        logger.info("[EXPORT] flavor=%s → fs_mode=%s", _flavor_key(util, fs_variant), cfg_flavor.get("fs_disc_mode"))

        logger.info("Coalition export (streaming) for flavor=%s", _flavor_key(util, fs_variant))
        payoff_map = write_from_cache(
            store,
            countries=countries,
            periods=periods,
            output_dir=output_dir,
            utility=util,
            L=L,
            tstep=tstep,
            base_year=base_year,
            config=cfg_flavor,
            selection=selection,
            backstop_switch_year=backstop_switch_year,
            export_selection=export_selection,
        )

        # If mega_run and we streamed ALL coalitions for this flavor, write the overview
        # even if coverage < 100%; the overview itself distinguishes certain vs maybe
        # stability and reports coverage explicitly.
        if bool(cfg.get("mega_run", False)) and (_spec in {"all", "*"}):
            expected_nonempty = (1 << len(countries)) - 1
            have = len([m for m in payoff_map.keys() if "1" in m])
            logger.info(
                "Building overview for flavor=%s (coverage %d/%d).",
                _flavor_key(util, fs_variant),
                have,
                expected_nonempty,
            )
            export_stability_overview_from_map(
                _flavor_key(util, fs_variant),
                payoff_map,
                countries=countries,
                periods=periods,
                output_dir=output_dir,
                config=cfg_flavor,
            )

def _attach_mu_plot(
    wb,
    solution: Dict[str, Any],
    countries: List[str],
    periods: List[int],
    *,
    base_year: int,
    tstep: int,
    backstop_switch_year: int | None = None,
    diagnostics_level: str = "minimal",
) -> None:
    """Create a 'mu Plot' sheet; respects diagnostics_level ('off' disables)."""
    if diagnostics_level == "off":
        return
    buf = BytesIO()
#    cap_year = None
#    if backstop_switch_year is not None:
#        cap_year = min(base_year + max(periods)*tstep, int(backstop_switch_year) + tstep)
    # alternative: hardcode 2125
    cap_year = 2125
    plot_mu_for_run(
        solution.get("mu", {}), countries, periods,
        base_year=base_year, tstep=tstep, year_cap=cap_year, outfile=buf
    )
    buf.seek(0)
    img = PILImage.open(buf)
    tmp = BytesIO(); img.save(tmp, format="PNG"); tmp.seek(0)
    ws = wb.create_sheet("mu Plot")
    ws.add_image(XLImage(tmp), "A1")
    if cap_year is not None:
        ws["A65"] = "x-axis limited (cap = backstop_switch_year + tstep)"



def _fetch_planner_solution_from_store(store: "CoalitionStore", countries: List[str], utility: str,
                                       neg_use: bool, neg_weights, T: int, disc_tag: str | None = None) -> dict | None:
    """Try exact spec-id first; fall back to scanning GRAND rows."""
    try:
        sid = build_solution_spec_id(
            utility=str(utility).lower(), T=int(T), countries=countries,
            population_weight_envy_guilt=(utility == "fs" and False),  # planner coop uses global weighting toggle elsewhere
            exogenous_S=None,
            negishi_use=bool(neg_use),
            negishi_weights=neg_weights,
            disc_tag=(str(disc_tag) if (utility == "fs") else None),
        )
        grand = tuple(1 for _ in countries)
        rec = store.get(grand, sid)
        sol = (rec or {}).get("solution")
        if sol: return sol
    except Exception:
        pass
    # Fallback: scan GRAND entries
    for r in store.iter_rows(spec_id=None):
        vec = tuple(r.vector) if r.vector is not None else None
        if vec != tuple(1 for _ in countries): 
            continue
        rec = store.get(vec, getattr(r, "spec_id", None))
        sol = (rec or {}).get("solution") or {}
        if str(sol.get("utility","")).lower() != str(utility).lower():
            continue
        if utility == "fs" and disc_tag:
            tag = str(sol.get("disc_tag","")).lower()
            if str(disc_tag).lower() not in tag:
                continue
        return sol or None
    return None